import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from logging_config import set_logging
from price_checker.checker_utilites import get_platform
from utilites import get_last_dir, check_dir

log = set_logging('price_checker')


class JsonPriceToXls:
    def __init__(self, actual_table):
        json_folder = 'price_checker/web_data/'
        self.date = get_last_dir(json_folder)
        self.src_folder = json_folder + self.date
        self.platform_prices = {}
        self.workbook = load_workbook(actual_table)

    def run(self):
        self.load_data()
        # for page in self.workbook.get_sheet_names():
        #     self.extend_table(page)
        self.fill_prices('rosel')
        self.fill_prices('oppo')
        self.check_conditions()
        self.write_result()

    def load_data(self):
        prices = self.platform_prices
        for json_filename in os.listdir(self.src_folder):
            platform, page = json_filename.split('.')[0].split('_')
            if not prices.get(page):
                prices[page] = {}
            with open(f'{self.src_folder}/{json_filename}', 'r', encoding='utf8') as read_file:
                data = json.load(read_file)
            prices[page][platform] = data

    def fill_prices(self, page):
        prices = self.platform_prices[page]
        ws = self.workbook[page]
        price_col = self.add_new_price_column(ws)
        for order, row in enumerate(ws, 1):
            platform = get_platform(row[0].value)
            if not platform:
                continue
            row_id = f'{order:03}'
            price = prices[platform].get(row_id)
            try:
                price = float(price)
            except Exception as e:
                pass
            row[price_col].value = price

    def write_result(self):
        filename = f'{check_dir("xls_results/price_ch")}/{self.date}_price_checker.xlsx'
        self.workbook.save(filename)
        log.info(f'{filename} saved')

    def add_new_price_column(self, ws):
        col_max = ws.max_column
        col_insert = col_max - 1
        col_oppo_price = col_max - 2
        ws.insert_cols(col_insert, amount=1)
        title_row = ws[1]
        ws.column_dimensions[title_row[col_insert].column_letter].width = 10
        ws.column_dimensions[title_row[col_oppo_price].column_letter].width = 10
        ws.column_dimensions[title_row[col_max].column_letter].width = 10
        new_date = title_row[col_oppo_price]
        new_date.value = self.date
        new_date.alignment = Alignment(horizontal='center', vertical='center')
        return col_oppo_price

    def check_conditions(self):
        self.page_conditions('rosel')
        self.page_conditions('oppo')

    def page_conditions(self, page):
        ws = self.workbook[page]
        max_column = ws.max_column - 1
        for row in ws:
            cell_rrc = row[max_column - 1]
            cell_oppo = row[max_column - 2]
            cell_oppo.number_format = '#,##0.00\\ "â‚½"'
            cell_oppo.alignment = Alignment(horizontal='right', vertical='center')
            rosel_price = row[max_column].value if row[max_column].value else cell_rrc.value
            if self.incorrect_price(rosel_price, cell_oppo):
                continue
            color_yes, color_no = 'C4D79B', 'E6B8B7'
            difference = cell_oppo.value - rosel_price
            if abs(difference) >= rosel_price / 10:
                fg_color = color_no if difference > 0 else color_yes
                cell_oppo.fill = PatternFill("solid", fgColor=fg_color)

    def incorrect_price(self, rosel_price, cell_oppo):
        oppo_price = cell_oppo.value
        if oppo_price == 'ERROR':
            cell_oppo.fill = PatternFill("solid", fgColor='ff0000')
            cell_oppo.alignment = Alignment(horizontal='center', vertical='center')
            return True
        if not rosel_price or not oppo_price:
            return True
        if type(oppo_price) == str or type(rosel_price) == str:
            return True
        return False
