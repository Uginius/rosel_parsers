import json
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
from openpyxl.styles.numbers import BUILTIN_FORMATS
from config import date_template, dir_template
from product_representation.src.prre_goals import oz_goals, wb_goals
from utilites import get_last_dir, check_dir

shoplist = ['oz', 'wb']
sellers = ['РОСЭЛ', 'ФОТОН', 'SAFELINE', 'КОНТАКТ', 'КОНТАКТ ДОМ', 'РЕКОРД', 'ORGANIDE']
goals_and_terms = {'oz': oz_goals, 'wb': wb_goals}


def last_month_json_files(shop):
    folder = "product_representation/json_files"
    folder = os.path.join(folder, get_last_dir(folder))
    jsf = {}
    for filename in os.listdir(folder):
        find_date = re.findall(dir_template, filename)
        is_platform = shop in filename
        if find_date and is_platform:
            jsf[datetime.strptime(find_date[0], date_template)] = filename
    dates = sorted(list(jsf.keys()))
    last = dates[-1]
    last_month_files = {d: f'{folder}/{jsf[d]}' for d in dates if d.year == last.year and d.month == last.month}
    return last_month_files


def get_first_rqs(platform):
    last_cat = ''
    first_requests = {}
    for rq, terms in goals_and_terms[platform].items():
        category = terms['category']
        if category != last_cat:
            first_requests[rq] = category
            last_cat = category
    return first_requests


class PrReMonthlyDataAnalytics:
    def __init__(self):
        self.current_goal = None
        self.platform_requests = {}
        self.platform_goods = []
        self.all_platform_goods = {}
        self.categories = {}
        self.date = {}
        self.rosel_goods = {}
        self.workbook = Workbook()
        self.shop = None
        self.sheet = None
        self.line_kpi = 0
        self.total_kpi = 0

    def run(self):
        self.initiate_workbook()
        self.platform_actions(shoplist[0])
        self.platform_actions(shoplist[1])
        self.save_table()

    def initiate_workbook(self):
        for shop in shoplist:
            self.workbook.create_sheet(shop)
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def platform_actions(self, platform):
        self.total_kpi, self.shop = 0, platform
        self.get_json()
        self.add_caption()
        self.add_titles()
        self.add_body()
        self.set_total_kpi()

    def get_json(self):
        platform = self.shop
        json_filenames = {'oz': last_month_json_files('oz'), 'wb': last_month_json_files('wb')}
        goods = {}
        for date, filename in json_filenames[platform].items():
            print(f'opening {filename}')
            with open(filename, 'r', encoding='utf8') as file:
                json_req = dict(json.load(file))
            self.set_rosel_goods(json_req)
            goods[date] = self.rosel_goods
        self.all_platform_goods[platform] = goods

    def set_rosel_goods(self, json_data):
        self.rosel_goods = {}
        for request_id, search_results in json_data.items():
            true_positions = {}
            for order in search_results:
                merch = search_results[order]
                loaded_brand = merch['brand']
                if not loaded_brand:
                    continue
                if loaded_brand.upper() in sellers:
                    merch_name = merch['name']
                    merch_id = merch['shop_id']
                    true_positions[int(order)] = f'{merch_id}, №{order}: {merch_name}'
            self.rosel_goods[request_id] = true_positions

    def check_incorrect_goods(self, req_id, prod_id):
        ban_prs = {'индикаторная отвертка': ['38710001', '79676933'], 'отвертка индикаторная': ['38710001', '79676933']}
        incorrect_product_id_detected = False
        if self.platform_requests[req_id] in ban_prs.keys():
            ban_list = ban_prs[self.platform_requests.get(req_id, None)]
            if prod_id in ban_list:
                incorrect_product_id_detected = True
        return incorrect_product_id_detected

    def add_caption(self):
        first_line = ['Отчет о представленности продукции по запросам на маркет-плейсах']
        platform = self.shop
        regions = {'oz': 'Москва', 'wb': 'Санкт-Петербург'}
        second_line = [f'Регион для которого осуществляется поисковая выдача - {regions[platform]}']
        self.sheet = self.workbook[platform]
        sw = self.sheet
        sw.append(first_line)
        sw.append([])
        sw.append(second_line)
        sw.append([])
        sw['A1'].font = Font(name='Calibri', size=20, color="000000", bold=True)
        sw.row_dimensions[1].height = 40
        sw.column_dimensions['A'].width = 35
        sw.column_dimensions['B'].width = 35
        sw.column_dimensions['C'].width = 7

    def add_titles(self):
        dates = [date.strftime('%d.%m.%Y') for date in self.all_platform_goods[self.shop]]
        products = ['Товары ' + date for date in dates]
        titles = ['Целевой запрос', 'Цель', 'КПД', *dates, *products]
        self.sheet.append(titles)
        self.add_titles_style(row_num=self.sheet.max_row, dates_quantity=len(dates))

    def add_titles_style(self, row_num, dates_quantity):
        sw = self.sheet
        thick = Side(border_style="thick", color="000000")
        first_date_column_number = ord('D')
        first_goods_column_number = first_date_column_number + dates_quantity
        for n in range(dates_quantity):
            date_column = chr(first_date_column_number + n)
            date_cell = sw[f'{date_column}{row_num}']
            date_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
            date_column_dim = sw.column_dimensions[date_column]
            date_column_dim.width = 10
            date_column_dim.alignment = Alignment(horizontal='center')
            goods_column = chr(first_goods_column_number + n)
            goods_column_dim = sw.column_dimensions[goods_column]
            goods_column_dim.width = 50

    def add_body(self):
        platform = self.shop
        self.platform_goods = [date for date in self.all_platform_goods[platform].values()]
        goals = goals_and_terms[platform]
        first_rqs = get_first_rqs(platform)
        for req_id, terms in goals.items():
            if first_rqs.get(req_id):
                self.check_category(terms['category'])
                continue
            self.current_goal = terms['goal']
            self.set_body_lines(req_id, terms)

    def set_body_lines(self, req_id, terms):
        self.line_kpi = 0
        sw = self.sheet
        first_row_number = sw.max_row + 1
        all_dates_shop_products_for_rq = [prods[req_id] for prods in self.platform_goods]
        quantity_in_dates = [len(el) for el in all_dates_shop_products_for_rq]
        max_goods_q = max(quantity_in_dates)
        products_in_dates = [prs for prs in all_dates_shop_products_for_rq]
        if max_goods_q:
            self.body_lines_to_table(quantity_in_dates=quantity_in_dates, products_in_dates=products_in_dates,
                                     max_goods_q=max_goods_q, terms=terms)
        else:
            empty = [None] * len(quantity_in_dates)
            kpi = None
            result = [terms['request'], self.current_goal, kpi, *quantity_in_dates, *empty]
            sw.append(result)
        self.merge_body_cells(fst_row=first_row_number, lens=quantity_in_dates, req_id=req_id)
        # self.row_top_borders(current_row_number)

    def body_lines_to_table(self, quantity_in_dates, products_in_dates, max_goods_q, terms):
        dates_amount = len(quantity_in_dates)
        dates_lines = []
        for date in products_in_dates:
            date_products = list(date.values())
            len_date = len(date_products)
            if len_date < max_goods_q:
                for num in range(len_date, max_goods_q):
                    date_products.append(None)
            dates_lines.append(date_products)
        for prod_number in range(max_goods_q):
            products_count = quantity_in_dates if prod_number == 0 else [None] * dates_amount
            products_line = [dates_lines[date][prod_number] for date in range(dates_amount)]
            line = [terms['request'], terms['goal'], None, *products_count, *products_line]
            self.sheet.append(line)

    def merge_body_cells(self, fst_row, lens, req_id):
        max_q = max(lens)
        last_row = fst_row + max_q - 1 if max_q else fst_row
        # last_row = fst_row + max_q - 1
        self.merge_abc(fst=fst_row, last=last_row)
        self.merge_goods_cells(first_row=fst_row, lens=lens, req_id=req_id)

    def merge_abc(self, fst, last):
        sw = self.sheet
        sw[f'A{fst}'].alignment = Alignment(horizontal='left', vertical='center')
        sw[f'B{fst}'].alignment = Alignment(horizontal='left', vertical='center')
        sw[f'C{fst}'].alignment = Alignment(horizontal='center', vertical='center')
        if last:
            sw.merge_cells(f'A{fst}:A{last}')
            sw.merge_cells(f'B{fst}:B{last}')
            sw.merge_cells(f'C{fst}:C{last}')

    def merge_goods_cells(self, first_row, lens, req_id):
        sw = self.sheet
        first_cell = ord('D')
        rows = max(lens)
        for n, goods_in_page in enumerate(lens):  # 4 0 5 0
            column = chr(first_cell + n)
            cell = f'{column}{first_row}'
            if goods_in_page:
                last_row = first_row + goods_in_page - 1
                sw.merge_cells(f'{cell}:{column}{last_row}')
            else:
                last_row = first_row + rows - 1 if rows else first_row
                # last_row = first_row + rows - 1
                sw.merge_cells(f'{cell}:{column}{last_row}')
                column2 = chr(first_cell + n + len(lens))
                cell2 = f'{column2}{first_row}'
                sw.merge_cells(f'{cell2}:{column2}{last_row}')
            sw[cell].alignment = Alignment(horizontal='center', vertical='center')
            self.check_goals(cell, self.platform_goods[n][req_id])
            self.row_top_borders(first_row)
        kpi_cell = sw[f'C{first_row}']
        kpi_cell.value = self.line_kpi

    def check_goals(self, cl, goods):
        cell = self.sheet[cl]
        cond = self.is_conditions_true(cell=cell, goods=goods)
        color = 'C4D79B' if cond else 'E6B8B7'
        cell.fill = PatternFill("solid", fgColor=color)
        self.total_kpi += 1
        if cond:
            self.line_kpi += 1

    def row_top_borders(self, row_number):
        thin = Side(border_style="thin", color="000000")
        for cell in self.sheet[row_number]:
            cell.border = Border(top=thin)

    def is_conditions_true(self, cell, goods):
        match self.current_goal:
            case 'на 1 странице, по популярности':
                return True if cell.value > 0 else False
            case 'не менее 5 SKU на 1 странице, по популярности':
                return True if cell.value >= 5 else False
            case 'не менее 4 SKU на 1 странице, по популярности':
                return True if cell.value >= 4 else False
            case 'не менее 2 SKU на 1 странице, по популярности':
                return True if cell.value >= 2 else False
            case 'не ниже 5 строки, по популярности':
                top5 = []
                for prod in goods:
                    if prod:
                        if prod < 6:
                            top5.append(prod)
                # top5 = [prod.order for prod in goods if prod.order < 6]
                return True if top5 else False

    def check_category(self, category_name):
        self.sheet.append([category_name])
        max_row = self.sheet.max_row
        max_column = chr(ord('A') + self.sheet.max_column - 1)
        self.sheet.merge_cells(f'A{max_row}:{max_column}{max_row}')
        self.set_style_to_categoy_cell(max_row)

    def set_style_to_categoy_cell(self, row_order):
        cat_font = Font(name='Calibri', size=11, color="000000", bold=True)
        thin = Side(border_style="thin", color="000000")
        category_cell = self.sheet[f'A{row_order}']
        category_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        category_cell.fill = PatternFill("solid", fgColor="a6a6a6")
        category_cell.font = cat_font
        rd = self.sheet.row_dimensions[row_order]
        rd.height = 15

    def set_total_kpi(self):
        sw = self.sheet
        row = sw.max_row
        kpi_result = f'=SUM(C7:C{row})/{self.total_kpi}'
        sw.append([None, None, kpi_result])
        result_cell = sw[f'C{row + 1}']
        result_cell.number_format = BUILTIN_FORMATS[10]

    def save_table(self):
        last_date = list(self.all_platform_goods['oz'])[0].strftime('%Y-%B')
        folder = 'xls_results/prre'
        check_dir(folder)
        self.workbook.save(f'{folder}/representation_{last_date}.xlsx')
